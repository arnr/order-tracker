# All functions 

#from curses import window
#from msilib import datasizemask
from cgi import test
from cgitb import text
from tkinter import CENTER
from glob import escape
from numpy import can_cast
import openpyxl
import os
import pandas as pd
import numpy
import re
from tkinter import *
from tkinter import ttk
from functions import *
from tkinter import messagebox
from tkcalendar import Calendar
import sys
from functools import partial
import time

def new_order(worksheet, ticket_num, date, vendor, part_name, part_num, order_num, amount, pay_method ):
    # enters new data row into sheet

    ref_row = worksheet.max_row + 1
    data = [ticket_num, vendor, part_name, part_num, order_num, amount, pay_method]

    worksheet.cell(row=ref_row, column=1).value = 'ordered ' + date +';'

    for i in range(2,9):
        worksheet.cell(row = ref_row, column = i).value = data[i-2]
    
    return worksheet
    # TO DO:
    # this function may need to be rewritten due to using openpyxl. if changes are made the dataframe 
    # object will not have the data. essentially program will have to pull the sheet again. not optimal
   

def updateStatus(df, newStatus, date, rowNum):
    #ws.cell(row = rowNum, column = columnNum).value = newStatus + date +'; ' + ws.cell(row = rowNum, column = columnNum).value 
    if df.at[rowNum, 'CURRENT STATUS'] != None:
        df.at[rowNum, 'STATUS HISTORY'] = df.at[rowNum, 'CURRENT STATUS'] + '; ' + df.at[rowNum, 'STATUS HISTORY'] 

    df.at[rowNum, 'CURRENT STATUS'] = newStatus + ' ' + date


def search(df, colName, searchTerm,):
    #show rows based on a term 'searchTerm' down a particular column 'colName'
     result = df[df[colName].str.contains(searchTerm)]
     return result





def save(wb, filename):
    wb.save(filename)


def show(df):
    print(df)


## TO DO: use below function for reference. delete in final version
## help to select and edit rows data in treeview
def edit(tree):
   # Get selected item to Edit
   selected_item = tree.selection()[0]
   tree.item(selected_item, text="blub", values=("foo", "bar"))

## TO DO: delete in final code. just placed for ease of navigation in outline mode in vscode
def PLACE_HOLDER():
    escape



#deleting and editing rows in treeview: see below tutorial
#https://www.tutorialspoint.com/delete-and-edit-items-in-tkinter-treeview

## General functions
## 

def remove(string):
    return "".join(string.split())

    '''
    SOURCE
    https://www.geeksforgeeks.org/python-remove-spaces-from-a-string/
    '''

#---------------------------------------------------    



## Data related functions
##


# gets data from main excel sheet and puts it into a pandas dataframe. 
# Dataframe is processed to show only wanted data.
def load_excel():

    cwd = os.getcwd()
    file_directory = os.path.join(cwd, 'program', 'files')
    os.chdir(file_directory)
    wb = openpyxl.load_workbook('order tracking.xlsx')
    ws = wb['ORDERS']
        
    # setup pandas dataframe for further processing
    df = pd.DataFrame(ws.values)

    ## make sure that first row is actually the title row
    df.columns = df.iloc[0]
    df = df[1:]

    ## make sure these values treated as strings
    df['PART #'] = df['PART #'].astype(str) 
    df['ORDER #'] = df['ORDER #'].astype(str)

    return df


# show rows for formated for proper displaying
# if index_list is given then it will show only the rows corresponding to the number in index list
def show_entries(dataframe:pd.DataFrame, index_list=[]):
    
    if index_list != []:
        result = dataframe.loc[index_list].loc[:,dataframe.columns != 'STATUS HISTORY']
    else:
        result = dataframe.loc[:,dataframe.columns != 'STATUS HISTORY']
    return result


# gets status history. It can bring entire dataframe set or just one row info
def get_status_history(dataframe:pd.DataFrame, index=-1):
    
    # regex definition
    status_regex = re.compile(r'\w+\s\d+\/\d+\/\d\d\d\d')

    # get all status histories
    if index == -1:
        data = dataframe['STATUS HISTORY'].tolist()
        return data

    # get status history at index row
    else:
        data = dataframe.iloc[index].loc['STATUS HISTORY']
        
        try:
            list = status_regex.findall(data)
            
        except TypeError:
            list = []


        if list == []:
            messagebox.showinfo("No History", "This record has no history")

        return list
        
    '''
    SOURCE
    https://stackoverflow.com/questions/22341271/get-list-from-pandas-dataframe-column-or-row
    https://stackoverflow.com/questions/28754603/indexing-pandas-data-frames-integer-rows-named-columns
    '''


# function to save the edited treeview data back to dataframe for further saving
def saveable_dataframe(original_dataframe, tree):
    # origianl dataframe = dataframe that was loaded in to the treeview before any editing.
    # tree = treeview object with all the edited data. 

    # pull data from treeview widget and pass to dataframe.
    values = []
    for child in tree.get_children():
        values.append(tree.item(child)["values"])
    
    df_to_write = pd.DataFrame(values)
    
    # status history and add it back to dataframe
    list = get_status_history(original_dataframe)
    df_to_write.insert(1,"STATUS HISTORY", list, False)
    
    return df_to_write


# function to write the final data from a dataframe to the excel sheet
def write_to_excel(edited_dataframe:pd.DataFrame):
    #edited_dataframe: the dataframe with all the edits in it. The final data should be here.
   
    writer = pd.ExcelWriter('order tracking test.xlsx')
    edited_dataframe.to_excel(writer,sheet_name='ORDERS',index=False)
    writer.close()

#---------------------------------------------------



## General GUI related functions
##


def grey_out(window):
    # window is name of window/widget to "greyed out"
    
    try:
        for child in window.winfo_children():
            wtype = child.winfo_class()
            
            if wtype not in ('Frame','Labelframe'):
                child.configure(state='disable')
            
            else:
                grey_out(child)
   
    except TclError:
        pass


def grey_in(window):
    # window is name of window/widget to "greyed in"
    
    try:
        for child in window.winfo_children():
            wtype = child.winfo_class()
            
            if wtype not in ('Frame', 'Labelframe'):
                child.configure(state= 'normal')
            
            else:
                grey_in(child)
   
    except:
        pass
    '''
    SOURCE
    https://stackoverflow.com/questions/24942760/is-there-a-way-to-gray-out-disable-a-tkinter-frame
    https://www.tutorialspoint.com/getting-every-child-widget-of-a-tkinter-window
    https://www.tutorialspoint.com/how-to-gray-out-disable-a-tkinter-frame
    '''


# function for the close button
def window_close(parent_window, child_window):
    child_window.grab_release()
    grey_in(parent_window)
    child_window.destroy()


# creates a 3 second splash screen
def splash():
    splash_screen = Tk()
    splash_screen.geometry("300x200")
    
    splash_screen.overrideredirect(True) # removes the top window handle bar
    splash_screen.eval('tk::PlaceWindow . center') # center window to screen
    
    splash_label = Label(splash_screen, text="Splash text welcome")
    splash_label.place(relx=.5, rely=0.5, anchor='n')

    splash_screen.after(3000,splash_screen.destroy) # close window after 3 seconds
    '''
    SOURCE
    https://www.youtube.com/watch?v=LTVvHObxc4E 
    '''
    # TO DO: this is a basic splash screen. You need to put logo or picture or some decoration on it to finalize this section


# takes the treeview object (table) and populates it with data from provided dataframe (dataframe) 
def populate_table(table:ttk.Treeview, dataframe:pd.DataFrame): 

    #clear table of old data
    table.delete(*table.get_children())

    #defining the columns
    table["column"] = list(dataframe.columns)
    table["show"] = "headings"
        
    #formatting the columns
    for col in table['column']:
        table.column(col,anchor=CENTER, width=40)
        table.heading(col, text=col)
    
    #testing code
    for col in table['columns']:
        table.column(col, width=118)
    ## TO DO: please read
    #https://stackoverflow.com/questions/64477453/how-to-view-partial-area-of-a-treeview-and-horizontal-scrollable-in-tkinter


    #populate table with data from dataframe
    df_rows = dataframe.to_numpy().tolist()
    for row in df_rows:
        table.insert("","end", values=row)
    
    return table
    
#---------------------------------------------------



## GUI buttons functions
##

# level 2 windows hold secondary data or do seconday activities and are called from the main window. 
# main_window is the name of window from which this window will be called. usually it will be 'root' window
def open_lvl2_window(new_window, window_title:str, main_window:Tk, size="200x200"):

    new_window = Toplevel(main_window)
    new_window.title(window_title)
    new_window.geometry(size)

    return new_window


# level 3 window shows history of statuses of selected row
def status_history_window(window, dataframe, index):
    # window = name of window from where this window will get called
    # dataframe and list for the get_status_history function
    
    # get data for this window; exit function if none available
    list = get_status_history(dataframe, index)
    if list == []:
        return
     
    
    # draw window
    status_window = Toplevel(window)
    status_window.title("Status History")
    status_window.geometry("250x300")
    status_window.grab_set()

    # box for data 
    listbox = Listbox(status_window)    
    listbox.place(relx = 0.5, rely = 0.1, relwidth = 0.9, relheight=0.8, anchor=N)

    # button function
    # def window_close():
    #     status_window.grab_release()
    #     grey_in(window)
    #     status_window.destroy()
    
    # button
    close_btn = Button(status_window, text = 'Close', command=lambda: window_close(window, status_window))
    close_btn.place(relx=0.5, rely=0.85, relwidth=0.9, anchor=N)

    # populate box
    for count, item in enumerate(list):
       listbox.insert(count, item)

    # grey out the parent window
    grey_out(window)

# options for edit buttons in called throught the edit_row funciton
status_options = [
'ordered',
'reordered',
'received',
'ticket_compeleted',
'return_requested',
'return_authorized',
'credit_memo_received'
]


def get_date(parent_window):

    # draw window
    date_window = Toplevel(parent_window)
    date_window.title("Select Date")
    date_window.geometry("200x400")
    date_window.grab_set()

    # draw calender
    cal = Calendar(date_window, selectmode = 'day',
            year = 2020, month = 5, day = 22,
            date_pattern='MM/dd/yyyy')
    cal.pack(pady = 20)

    # holds the date (declared empty)
    get_date.selected_date=''


    # button functions

    # gets the date info from calender object to the necessary variables.
    def grab_date():
        dateinfo.config(text = "Selected Date is: " + cal.get_date())
        get_date.selected_date = cal.get_date()
        

    # Add Button and Label
    Button(date_window, text= "Get Date", command= grab_date).pack(pady = 5)

    dateinfo = Label(date_window, text = "")
    dateinfo.pack(pady = 10)

    close_btn = Button(date_window, text = 'Save & Close', command=lambda: window_close(parent_window, date_window))
    close_btn.pack(pady= 10)

    grey_out(parent_window)

    '''SOURCE
    https://www.geeksforgeeks.org/create-a-date-picker-calendar-tkinter/
    https://www.codesdope.com/blog/article/nested-function-scope-of-variable-closures-in-pyth/
    '''


def edit_status(parent_window, value):

    # Draw window
    edit_status_window = Toplevel(parent_window)
    edit_status_window.title("Edit Status")
    edit_status_window.geometry("200x250")
    edit_status_window.grab_set()        

    # Draw label
    label_status = Label(edit_status_window, bg= "white", relief= SUNKEN, anchor= "w" )
    label_status.place(relx = 0.5, rely = 0.1, relwidth = 0.9, anchor= N)
    label_status.configure(text=value)

    # drop down menu for status
    clicked = StringVar()
    clicked.set('Select Status')
    drop = OptionMenu( edit_status_window , clicked,*status_options )
    drop.place(relx = 0.5, rely = 0.25, relwidth = 0.91, anchor= N)


    # Button functions
    edit_status.new_status='nothing'

    # TO DO: check all the error messages actually trigger properly. was working but changes have rendered some of it not. I noticed in some of the runs
    def save(window):
        # validate if date and status were selected before accepting status entry
        try:
            if get_date.selected_date:
                if not clicked.get() == 'Select Status':
                    edit_status.new_status = clicked.get() + ' ' + get_date.selected_date
                    label_status.configure(text= edit_status.new_status)                
                    
                    # wait for selected info to show in status label 
                    time.sleep(2)
                    window.destroy()
                
                else:
                   messagebox.showerror("ERROR", "Status not selected!") 
            
            else:
                messagebox.showerror("ERROR", "Date not selected!")
        
        except AttributeError:
            messagebox.showerror("ERROR", "Date not selected!")

    # draw buttons
    date_btn = Button(edit_status_window, text='Date', command=lambda: get_date(edit_status_window))
    date_btn.place(relx = 0.5, rely = 0.38, relwidth = 0.9, anchor= N)

    save_btn = Button(edit_status_window, text='Save', command=lambda: save(edit_status_window))
    save_btn.place(relx = 0.5, rely = 0.55, relwidth = 0.9, anchor= N)
    
    # make the window wait till user puts data for rest of program to continure
    parent_window.wait_window(edit_status_window)

    '''SOURCE
    https://www.geeksforgeeks.org/dropdown-menus-tkinter/
    https://stackoverflow.com/questions/28388346/what-does-thewait-window-method-do
    '''


def edit_vendor(parent_window, value):
    
    # Draw window
    edit_vendor_window = Toplevel(parent_window)
    edit_vendor_window.title("Edit Vendor")
    edit_vendor_window.geometry("200x250")
    edit_vendor_window.grab_set()    

    # Draw label
    label_vendor = Label(edit_vendor_window, bg= "white", relief= SUNKEN, anchor= "w" )
    label_vendor.place(relx = 0.5, rely = 0.1, relwidth = 0.9, anchor= N)
    label_vendor.configure(text=value)

    vendor_list = open('vendor list.txt').read().splitlines()
    
    clicked = StringVar()
    clicked.set('Select Vendor')
    drop = OptionMenu(edit_vendor_window , clicked, *vendor_list)
    drop.place(relx = 0.5, rely = 0.25, relwidth = 0.91, anchor= N)

    def save(window):
        if not clicked.get() == 'Select Vendor':
            edit_vendor.new_vendor = clicked.get()
            label_vendor.configure(text= edit_vendor.new_vendor)                
                        
            # wait for selected info to show in status label 
            time.sleep(2)
            window.destroy()        

        else:
            messagebox.showerror("ERROR", "Vendor not selected")

    save_btn = Button(edit_vendor_window, text='Save', command=lambda: save(edit_vendor_window))
    save_btn.place(relx = 0.5, rely = 0.55, relwidth = 0.9, anchor= N)    

    # make the window wait till user puts data for rest of program to continure
    parent_window.wait_window(edit_vendor_window)


def edit_paymethod(parent_window, value):
    
    # Draw window
    edit_paymethod_window = Toplevel(parent_window)
    edit_paymethod_window.title("Edit Vendor")
    edit_paymethod_window.geometry("300x250")
    edit_paymethod_window.grab_set()    

    # Draw label
    label_paymethod = Label(edit_paymethod_window, bg= "white", relief= SUNKEN, anchor= "w" )
    label_paymethod.place(relx = 0.5, rely = 0.1, relwidth = 0.9, anchor= N)
    label_paymethod.configure(text=value)

    paymethod_list = open('payment method list.txt').read().splitlines()
    
    selected_method = StringVar()

    for a, method in enumerate(paymethod_list):
        rbutton = Radiobutton(
            edit_paymethod_window, 
            text= method, 
            variable=selected_method, 
            value=method, 
            command=lambda: label_paymethod.configure(text=selected_method.get())
            )
        
        if (a%2) == 0:
            rbutton.place(relx = 0.25, rely = 0.25 + (a/10), relwidth = 0.4, anchor= N)
        else:
            rbutton.place(relx = 0.75, rely = 0.25 + ((a-1)/10), relwidth = 0.4, anchor= N)


    def save():
    # clicked = StringVar()
    # clicked.set('Select Vendor')
    # drop = OptionMenu(edit_paymethod_window , clicked, *vendor_list)
    # drop.place(relx = 0.5, rely = 0.25, relwidth = 0.91, anchor= N)

    # def save(window):
    #     if not clicked.get() == 'Select Vendor':
    #         edit_paymethod.new_vendor = clicked.get()
    #         label_paymethod.configure(text= edit_paymethod.new_vendor)                
                        
    #         # wait for selected info to show in status label 
    #         time.sleep(2)
    #         window.destroy()        

    #     else:
    #         messagebox.showerror("ERROR", "Vendor not selected")
            
    # save_btn = Button(edit_paymethod_window, text='Save', command=lambda: save(edit_paymethod_window))
    # save_btn.place(relx = 0.5, rely = 0.55, relwidth = 0.9, anchor= N)    

    # make the window wait till user puts data for rest of program to continure
    grey_out(parent_window)
    parent_window.wait_window(edit_paymethod_window)

    '''SOURCE
    https://www.tutorialspoint.com/python/tk_radiobutton.htm
    https://www.pythontutorial.net/tkinter/tkinter-stringvar/
    '''


# level 2 window to show and edit one row from the treeview
def edit_row(tree, dataframe, parent_window):
        
    #retrieve data from selected line
    focused_line = tree.focus()
    focused_index = tree.index(tree.selection())
    '''
    SOURCE:
    https://stackoverflow.com/questions/68508694/how-do-i-get-the-index-of-selected-row-in-tkinter-treeview
    '''
    
    # check if a row was selected before moving on.
    if "".__eq__(focused_line):
        print("no data selected")
        messagebox.showerror("ERROR", "Nothing was selected to edit!")
        return
        '''
        SOURCE:
        https://stackoverflow.com/questions/9573244/how-to-check-if-the-string-is-empty
        '''    
     
    # GUI related
    #
    edit_window = Toplevel(parent_window)
    edit_window.title("Edit Entry")
    edit_window.geometry("350x400")
    edit_window.grab_set()

    '''CODE DEPRECATED - see note 1
    entry_boxes = {
        "status" : StringVar(),
        "ticket_num" : StringVar(), 
        "vendor" : StringVar(),
        "part_desc" : StringVar(),
        "part_num" : StringVar(),
        "order_num" : StringVar(),
        "amount" : StringVar(),
        "pay_method" : StringVar()
    }
    '''
    

    # store entry box data in these lists
    box_entries=[]
    label_text = ["Status", "Ticket #", "Vendor", "Part Description", "Part #", "Order #", "Amount", "Payment Method"]
    omit_list = ("Status", "Vendor", 'Payment Method') # so these can be made in to label box instead of entry boxes
    selected_data = tree.item(focused_line, "values")


    # main button & window functions 
    #

    # commits the changes typed in edit window back into the treeview
    def commit_changes():

        values=[]
        ''''''
        for entry in box_entries:
            if isinstance(entry, Entry):
                values.append(entry.get())
            # elif isinstance(entry, Label):
            #     values.append(omit_list_label.cget("text"))
        '''
        SOURCE
        https://blog.finxter.com/how-to-determine-the-type-of-an-object-in-python/
        https://stackoverflow.com/questions/34667710/pattern-matching-tkinter-child-widgets-winfo-children-to-determine-type
        '''
        # write to treeview
        tree.item(focused_line, values=values)

        # close window after saving
        window_close(parent_window, edit_window)   

    
    # function and list to deal with the edit buttons for the labels in edit screen
    label_btn_id = [] # reference list to buttons created in loop associated with omit list
    entry_bx_id = [] # reference list to all entry boxes for later calling
    

    def call_edit_fnct(i):
        # call for status edit button
        if i == 0:
            # call function to get changed value
            edit_status(edit_window, box_entries[i].get())

            # display new value in entry box
            entry_bx_id[i].config(state='normal')
            entry_bx_id[i].delete(0, END)
            entry_bx_id[i].insert(0,edit_status.new_status)
            entry_bx_id[i].config(state="disabled")
        
        # call for vendor edit button
        elif i == 2:
            # call function to get changed value
            edit_vendor(edit_window, box_entries[i].get())

            # display new value in entry box
            entry_bx_id[i].config(state='normal')
            entry_bx_id[i].delete(0, END)
            entry_bx_id[i].insert(0,edit_vendor.new_vendor)
            entry_bx_id[i].config(state="disabled")

        # call for payment method edit button        
        elif i == 7:
            # call function to get changed value
            edit_paymethod(edit_window, box_entries[i].get())

            # display new value in entry box
            entry_bx_id[i].config(state='normal')
            entry_bx_id[i].delete(0, END)
            entry_bx_id[i].insert(0,edit_vendor.new_vendor)
            entry_bx_id[i].config(state="disabled")

        # bname.configure(text = "clicked")
        # print(i)
    
    
    # create the labels and entry boxes; populate with relevant data
    for i in range(len(label_text)):
        
        # labels
        label = Label(edit_window)
        label.place(relx = 0.03, rely = (i/10) + 0.025, relwidth = 0.33)
        label.configure(text=label_text[i])

        
        # TO DO: the label must be left aligned to make it looke better

        '''CODE DEPRECATED - see note 1
        entry boxes
        entry_box = key + "_entrybx" 
        entry_box = Entry(edit_window, textvariable=entry_boxes[key])
        entry_box.place(relx = 0.5, rely = (i/10) + 0.1, relwidth = 0.4)
        entry_box.insert(0, selected_data[i])
        '''

        # editable entry boxes 
        if label_text[i] not in omit_list:
            # draw and populate entry boxes
            entry_box = Entry(edit_window)
            entry_box.place(relx = 0.35, rely = (i/10) + 0.025, relwidth = 0.4)
            entry_box.insert(0, selected_data[i])
            entry_bx_id.append(entry_box)
            
            # save data into list
            box_entries.append(entry_box)
            label_btn_id.append("") # placeholder strings added to so numbering reference is correct   

        # unclicable label boxes to show data and corresponding edit buttons
        if label_text[i] in omit_list:
            # draw and populate labels
            # omit_list_label = Label(edit_window, bg= "white", relief= SUNKEN, anchor= "w" )
            # omit_list_label.place(relx = 0.35, rely = (i/10) + 0.025, relwidth = 0.4)
            # omit_list_label.configure(text= selected_data[i])
            
            entry_box = Entry(edit_window)
            entry_box.place(relx = 0.35, rely = (i/10) + 0.025, relwidth = 0.4)
            entry_box.insert(0, selected_data[i])
            entry_bx_id.append(entry_box)
            entry_box.config(state="disabled")

            ''' CODE DEPRECATED -- solution to get function reference by using a string
            # get the functions for the edit button. Functions names are 'edit_Status', 'edit_status' & 'edit_PaymentMethod'
            edit_function = locals().get('edit_' + remove(label_text[i]))
            '''

            # draw edit button
            label_edit_btn = Button(edit_window, text= 'Edit', command=partial(call_edit_fnct, i) )
            label_edit_btn.place(relx = 0.8, rely = (i/10) + 0.025, relwidth = 0.15)
            label_btn_id.append(label_edit_btn)
      
            # save data into list
            # box_entries.append(lambda: omit_list_label)
            box_entries.append(entry_box)

        '''SOURCE: 
        -> Default text in entry widget:  https://www.geeksforgeeks.org/how-to-set-the-default-text-of-tkinter-entry-widget/  
        -> create entry boxes with loop and retrieve data from boxes:  https://www.youtube.com/watch?v=H3Cjtm6NuaQ&t
        -> https://stackoverflow.com/questions/21495367/how-to-align-text-to-the-left
        -> Tkinter Relief styles:  https://www.tutorialspoint.com/python/tk_relief.htm
        -> Assigning functions to buttons that were created with loop: https://stackoverflow.com/questions/39447138/how-can-i-identify-buttons-created-in-a-loop
        -> https://stackoverflow.com/questions/10865116/tkinter-creating-buttons-in-for-loop-passing-command-arguments
        -> https://www.tutorialspoint.com/how-to-clear-the-entry-widget-after-a-button-is-pressed-in-tkinter
        note 1
        DEPRECATED CODE 
        Used the below tutorial to write the code as one approach to solve 
        writing multiple widgets with loop and then retrieving the info thereafter. 

        To create multiple widgets and reference them:
        https://www.youtube.com/watch?v=XerT3-rrOmQ
        END OF note 1
        '''    

   # TO DO: possible solution to retrieving data https://www.youtube.com/watch?v=H3Cjtm6NuaQ
        # '''
        # https://www.tutorialspoint.com/delete-and-edit-items-in-tkinter-treeview
        # '''
    
    
    # button definitions
    #
    
    save_edit_btn = Button(edit_window, text= 'Save Changes', command=commit_changes)
    save_edit_btn.place(relx= 0.25, rely= 0.8, relwidth= 0.4, anchor= N)

    status_history_btn = Button(edit_window, text= ' Status History', command=lambda: status_history_window(edit_window, dataframe,focused_index))
    status_history_btn.place(relx= 0.75, rely= 0.8, relwidth= 0.4, anchor= N)

    close_btn = Button(edit_window, text= 'Close', command=lambda: window_close(parent_window, edit_window))
    close_btn.place(relx= 0.5, rely= 0.9, relwidth= 0.9, anchor= N)
    '''
    SOURCE:
    https://www.geeksforgeeks.org/how-to-close-a-window-in-tkinter/
    '''

    grey_out(parent_window)
#---------------------------------------------------

def update_status(window):
    # window is the parent window.

    # draw window
    edit_status_window = Toplevel(window)
    edit_status_window.title("Edit Status")
    edit_status_window.geometry("250x300")
    #edit_status_window.grab_set()

    # button functions

    # button definitions
    #close_btn = Button(edit_status_window, text= 'Close', command=lambda: window_close(window, edit_status_window))
    #close_btn.place(relx= 0.5, rely= 0.9, relwidth= 0.9, anchor= N)


    # grey out parent window
    #grey_out(window)


    pass

'''
to take data from treeview to dataframe 
https://stackoverflow.com/questions/56898437/method-to-get-treeview-table-into-new-dataframe


to edit and delete items in treeview
https://www.tutorialspoint.com/delete-and-edit-items-in-tkinter-treeview



'''
# def backup():
#    #this function will create a backup of the main excel sheet before any changes are made. 
#     #Furthermore it will check how many backups are there and delete the oldest ones so as to control number of copies kept